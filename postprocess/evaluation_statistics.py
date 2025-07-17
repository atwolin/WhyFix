import re
import os
import json
import openpyxl
import pandas as pd
import numpy as np
from scipy.stats import ttest_rel, chisquare
from sklearn.metrics import cohen_kappa_score
from itertools import combinations


def restore_method_headers(source_headers_file, target_numbered_file, output_file):
    """
    使用來源檔案作為參照，還原目標 Excel 檔案中的所有區段標頭，並保持原始版面配置。

    # 運作原理:
      它使用一個仍然保留原始標頭的對應 Excel 檔案作為「地圖」。
      它會讀取來源檔案 A 欄的所有儲存格內容 (包含標頭及空白列)，
      並將它們寫入目標檔案，以確保版面結構完全一致。

    Args:
        source_headers_file (str): 包含正確、描述性標頭的 .xlsx 檔案路徑 (例如 'baseline')。
        target_numbered_file (str): 標頭以數字表示的 .xlsx 檔案路徑 (例如 '1', '2')。
        output_file (str): 還原後的 .xlsx 檔案將被儲存的路徑。
    """
    print("--- 開始標頭還原流程 ---")

    # --- 1. 驗證輸入檔案是否存在 ---
    if not os.path.exists(source_headers_file):
        print(f"錯誤：在 '{source_headers_file}' 找不到來源檔案")
        return
    if not os.path.exists(target_numbered_file):
        print(f"錯誤：在 '{target_numbered_file}' 找不到目標檔案")
        return

    print(f"來源 (地圖) 檔案: '{source_headers_file}'")
    print(f"目標 (數字) 檔案: '{target_numbered_file}'")

    # --- 2. 載入來源和目標工作簿 ---
    # 我們使用 openpyxl 以保留所有現有的儲存格格式。
    source_wb = openpyxl.load_workbook(source_headers_file)
    target_wb = openpyxl.load_workbook(target_numbered_file)

    # --- 3. 遍歷工作簿中的每個工作表 ---
    sheet_names = source_wb.sheetnames
    for sheet_name in sheet_names:
        print(f"  - 正在處理工作表: '{sheet_name}'...")

        if sheet_name not in target_wb.sheetnames:
            print(f"    - 警告: 在目標檔案中找不到工作表 '{sheet_name}'，已跳過。")
            continue

        source_ws = source_wb[sheet_name]
        target_ws = target_wb[sheet_name]

        # --- 4. 從來源檔案建立標頭地圖 ---
        headers_map = []
        for row_index in range(3, source_ws.max_row + 1, 2):
            header_cell = source_ws.cell(row=row_index, column=1)
            headers_map.append(header_cell.value)

        # --- 5. 將還原的標頭寫入目標檔案 ---
        for i, header_text in enumerate(headers_map):
            row_to_update = 3 + (i * 2)
            cell_to_update = target_ws.cell(row=row_to_update, column=1)
            cell_to_update.value = header_text

    # --- 6. 將修改後的工作簿儲存到新檔案 ---
    target_wb.save(output_file)
    print(f"\n✅ 成功還原標頭並儲存至 '{output_file}'")


def analyze_evaluation_results(restored_files_folder):
    """
    讀取並解析所有還原後的評估者檔案，分離出「方法」與「Ablation」區塊的原始資料。
    這個函式只負責資料提取，不進行統計計算。

    Args:
        restored_files_folder (str): 包含已還原 Excel 檔案的資料夾路徑。

    Returns:
        tuple: 包含兩個列表 (all_method_data, all_ablation_data)。
    """
    print("--- 開始解析評估檔案 ---")

    all_method_data = []
    all_ablation_data = []
    restored_files = [f for f in os.listdir(restored_files_folder) if f.endswith(".xlsx") and not f.startswith("~")]

    print(f"在 '{restored_files_folder}' 中找到 {len(restored_files)} 個檔案進行分析。")

    # --- 1. 遍歷每個還原後的檔案 ---
    for filename in restored_files:
        filepath = os.path.join(restored_files_folder, filename)
        evaluator_name = os.path.splitext(filename)[0].split('_')[-1].split('.')[0]
        print(f"  - 正在處理檔案: {filename} (評估者: {evaluator_name})")
        xls = pd.ExcelFile(filepath)

        # --- 2. 遍歷檔案中的每個工作表並分離資料 ---
        for sheet_name in xls.sheet_names:
            df_full_sheet = pd.read_excel(xls, sheet_name=sheet_name, header=None, dtype=str)

            key_sentence_for_categorization = df_full_sheet.iloc[0, 0]

            try:
                ablation_header_row_index = df_full_sheet[df_full_sheet[0] == 'Ablation'].index[0]
            except IndexError:
                print(f"    - 警告: 在工作表 '{sheet_name}' 中找不到 'Ablation' 標頭，已跳過。")
                continue

            dataset_type = 'fce' if 'fce' in sheet_name.lower() else 'longman'

            # 分離「方法」資料 (上半部)
            method_cols = ['Method', 'Content', 'Accuracy', 'Helpfulness']
            df_method = df_full_sheet.iloc[2:ablation_header_row_index-1, :len(method_cols)].copy()
            df_method.columns = method_cols
            df_method = df_method.dropna(subset=['Method'])
            df_method['evaluator'] = evaluator_name
            df_method['dataset'] = dataset_type
            df_method['key_sentence'] = key_sentence_for_categorization
            all_method_data.append(df_method)

            # 分離「Ablation」資料 (下半部)
            search_start_row = ablation_header_row_index + 1

            try:
                end_of_ablation_section_row = df_full_sheet[df_full_sheet[0] == 'Crucial Information Preservation'].index[0]
            except IndexError:
                end_of_ablation_section_row = len(df_full_sheet)

            ablation_cols = ['Ablation', 'Content', 'Accuracy', 'Helpfulness', 'Preference', 'Crucial Information Preservation']
            df_ablation = df_full_sheet.iloc[search_start_row:end_of_ablation_section_row, :len(ablation_cols)].copy()

            if df_ablation.empty:
                continue

            df_ablation.columns = ablation_cols
            df_ablation.dropna(subset=['Ablation'], inplace=True)

            df_ablation_filtered = df_ablation.copy()
            df_ablation_filtered['evaluator'] = evaluator_name
            df_ablation_filtered['dataset'] = dataset_type
            df_ablation_filtered['key_sentence'] = key_sentence_for_categorization
            all_ablation_data.append(df_ablation_filtered)

        # Save
        with open(os.path.join(restored_files_folder, 'method_data.json'), 'w') as file:
            pd.concat(all_method_data, ignore_index=True).to_json(file, orient='records', force_ascii=False)
        with open(os.path.join(restored_files_folder, 'ablation_data.json'), 'w') as file:
            pd.concat(all_ablation_data, ignore_index=True).to_json(file, orient='records', force_ascii=False)

    print("--- 所有檔案解析完成 ---")
    return all_method_data, all_ablation_data


def generate_summary_report(all_method_data, all_ablation_data, output_filepath):
    """
    根據提供的原始資料，計算統計數據 (包含總計)，並產生 Excel 報告。
    [**新功能**] 在 'Total' 列中，2, 1, 0 的計數改為顯示「平均數量」。
    """
    print("--- 開始產生統計報告 ---")

    # --- 1. 處理與分析「方法」資料 ---
    if not all_method_data:
        print("未找到可分析的『方法』資料。")
        stats_method = pd.DataFrame()
    else:
        master_method_df = pd.concat(all_method_data, ignore_index=True)
        fce_original_condition = (master_method_df['dataset'] == 'fce') & (master_method_df['Method'] == 'original')
        master_method_df = master_method_df[~fce_original_condition]
        master_method_df['Accuracy'] = pd.to_numeric(master_method_df['Accuracy'], errors='coerce')
        master_method_df['Helpfulness'] = pd.to_numeric(master_method_df['Helpfulness'], errors='coerce')

        num_evaluators = master_method_df['evaluator'].nunique()

        # 計算每個評估者在各個 dataset (fce/longman) 的統計數據
        stats_method = master_method_df.groupby(['evaluator', 'dataset', 'Method']).agg({
            'Accuracy': ['count', 'mean', 'std', ('2', lambda x: x.eq(2).sum()), ('1', lambda x: x.eq(1).sum()), ('0', lambda x: x.eq(0).sum())],
            'Helpfulness': ['count', 'mean', 'std', ('2', lambda x: x.eq(2).sum()), ('1', lambda x: x.eq(1).sum()), ('0', lambda x: x.eq(0).sum())]
        }).fillna(0)
        for col in ['Accuracy', 'Helpfulness']:
            for score in ['2', '1', '0']:
                stats_method[(col, score)] = stats_method[(col, score)].astype(int)
        stats_method = stats_method.round(5)

        # 計算每個評估者合併後的 (Combined) 統計數據
        per_evaluator_combined_stats = master_method_df.groupby(['evaluator', 'Method']).agg({
            'Accuracy': ['count', 'mean', 'std', ('2', lambda x: x.eq(2).sum()), ('1', lambda x: x.eq(1).sum()), ('0', lambda x: x.eq(0).sum())],
            'Helpfulness': ['count', 'mean', 'std', ('2', lambda x: x.eq(2).sum()), ('1', lambda x: x.eq(1).sum()), ('0', lambda x: x.eq(0).sum())]
        }).fillna(0)
        for col in ['Accuracy', 'Helpfulness']:
            for score in ['2', '1', '0']:
                per_evaluator_combined_stats[(col, score)] = per_evaluator_combined_stats[(col, score)].astype(int)
        per_evaluator_combined_stats = per_evaluator_combined_stats.round(5)
        per_evaluator_combined_stats['dataset'] = 'Combined'
        per_evaluator_combined_stats = per_evaluator_combined_stats.reset_index().set_index(['evaluator', 'dataset', 'Method'])

        # 計算每個資料集(fce/longman)的總計
        total_stats_from_master = master_method_df.groupby(['dataset', 'Method']).agg({
            'Accuracy': ['mean', 'std', ('2', lambda x: x.eq(2).sum()), ('1', lambda x: x.eq(1).sum()), ('0', lambda x: x.eq(0).sum())],
            'Helpfulness': ['mean', 'std', ('2', lambda x: x.eq(2).sum()), ('1', lambda x: x.eq(1).sum()), ('0', lambda x: x.eq(0).sum())]
        }).fillna(0)
        # [**修改**] 將總計數除以評估者數量，得到平均計數
        if num_evaluators > 0:
            for col in ['Accuracy', 'Helpfulness']:
                for score in ['2', '1', '0']:
                    total_stats_from_master[(col, score)] = total_stats_from_master[(col, score)] / num_evaluators
        total_stats_from_master = total_stats_from_master.round(5)

        per_evaluator_counts = stats_method[[('Accuracy', 'count'), ('Helpfulness', 'count')]]
        total_counts = per_evaluator_counts.groupby(['dataset', 'Method']).first()
        total_stats_method = pd.merge(total_counts, total_stats_from_master, on=['dataset', 'Method'], how='outer')
        total_stats_method = total_stats_method[[
            ('Accuracy', 'count'), ('Accuracy', 'mean'), ('Accuracy', 'std'), ('Accuracy', '2'), ('Accuracy', '1'), ('Accuracy', '0'),
            ('Helpfulness', 'count'), ('Helpfulness', 'mean'), ('Helpfulness', 'std'), ('Helpfulness', '2'), ('Helpfulness', '1'), ('Helpfulness', '0')
        ]]
        total_stats_method['evaluator'] = 'Total'
        total_stats_method = total_stats_method.reset_index().set_index(['evaluator', 'dataset', 'Method'])

        # 計算 FCE 與 Longman 合併(Combined)的總結果
        overall_total_stats_from_master = master_method_df.groupby('Method').agg({
            'Accuracy': ['mean', 'std', ('2', lambda x: x.eq(2).sum()), ('1', lambda x: x.eq(1).sum()), ('0', lambda x: x.eq(0).sum())],
            'Helpfulness': ['mean', 'std', ('2', lambda x: x.eq(2).sum()), ('1', lambda x: x.eq(1).sum()), ('0', lambda x: x.eq(0).sum())]
        }).fillna(0)
        # [**修改**] 將總計數除以評估者數量，得到平均計數
        if num_evaluators > 0:
            for col in ['Accuracy', 'Helpfulness']:
                for score in ['2', '1', '0']:
                    overall_total_stats_from_master[(col, score)] = overall_total_stats_from_master[(col, score)] / num_evaluators
        overall_total_stats_from_master = overall_total_stats_from_master.round(5)

        if not total_stats_method.empty:
            overall_total_counts = total_stats_method.groupby('Method')[[('Accuracy', 'count'), ('Helpfulness', 'count')]].sum()
        else:
            overall_total_counts = total_counts.groupby('Method').first()

        overall_total_stats = pd.merge(overall_total_counts, overall_total_stats_from_master, on='Method', how='outer')
        overall_total_stats = overall_total_stats[[
            ('Accuracy', 'count'), ('Accuracy', 'mean'), ('Accuracy', 'std'), ('Accuracy', '2'), ('Accuracy', '1'), ('Accuracy', '0'),
            ('Helpfulness', 'count'), ('Helpfulness', 'mean'), ('Helpfulness', 'std'), ('Helpfulness', '2'), ('Helpfulness', '1'), ('Helpfulness', '0')
        ]]
        overall_total_stats['evaluator'] = 'Total'
        overall_total_stats['dataset'] = 'Combined'
        overall_total_stats = overall_total_stats.reset_index().set_index(['evaluator', 'dataset', 'Method'])

        stats_method = pd.concat([stats_method, per_evaluator_combined_stats, total_stats_method, overall_total_stats])
        stats_method.sort_index(inplace=True)

    # --- 2. 處理與分析「Ablation」資料 ---
    if not all_ablation_data:
        print("未找到可分析的『Ablation』資料。")
        null_counts_table = pd.DataFrame()
        stats_ablation = pd.DataFrame()
    else:
        master_ablation_df = pd.concat(all_ablation_data, ignore_index=True)
        master_ablation_df['Content'] = master_ablation_df['Content'].replace(r'^\s*(\[\])?\s*$', np.nan, regex=True)

        null_counts = master_ablation_df['Content'].isnull().groupby(
            [master_ablation_df['evaluator'], master_ablation_df['dataset'], master_ablation_df['Ablation']]
        ).sum().astype(int)
        null_counts_table = null_counts.unstack(fill_value=0)

        if not null_counts_table.empty:
            total_nulls = null_counts_table.sum().to_frame('Total').T
            null_counts_table = pd.concat([null_counts_table, total_nulls])

        non_null_ablation_df = master_ablation_df.dropna(subset=['Content']).copy()
        if not non_null_ablation_df.empty:
            non_null_ablation_df['Accuracy'] = pd.to_numeric(non_null_ablation_df['Accuracy'], errors='coerce')
            non_null_ablation_df['Helpfulness'] = pd.to_numeric(non_null_ablation_df['Helpfulness'], errors='coerce')

            num_evaluators_ablation = non_null_ablation_df['evaluator'].nunique()

            stats_ablation = non_null_ablation_df.groupby(['evaluator', 'dataset', 'Ablation']).agg({
                'Accuracy': ['count', 'mean', 'std', ('2', lambda x: x.eq(2).sum()), ('1', lambda x: x.eq(1).sum()), ('0', lambda x: x.eq(0).sum())],
                'Helpfulness': ['count', 'mean', 'std', ('2', lambda x: x.eq(2).sum()), ('1', lambda x: x.eq(1).sum()), ('0', lambda x: x.eq(0).sum())]
            }).fillna(0)
            for col in ['Accuracy', 'Helpfulness']:
                for score in ['2', '1', '0']:
                    stats_ablation[(col, score)] = stats_ablation[(col, score)].astype(int)
            stats_ablation = stats_ablation.round(5)

            per_evaluator_ablation_combined = non_null_ablation_df.groupby(['evaluator', 'Ablation']).agg({
                'Accuracy': ['count', 'mean', 'std', ('2', lambda x: x.eq(2).sum()), ('1', lambda x: x.eq(1).sum()), ('0', lambda x: x.eq(0).sum())],
                'Helpfulness': ['count', 'mean', 'std', ('2', lambda x: x.eq(2).sum()), ('1', lambda x: x.eq(1).sum()), ('0', lambda x: x.eq(0).sum())]
            }).fillna(0)
            for col in ['Accuracy', 'Helpfulness']:
                for score in ['2', '1', '0']:
                    per_evaluator_ablation_combined[(col, score)] = per_evaluator_ablation_combined[(col, score)].astype(int)
            per_evaluator_ablation_combined = per_evaluator_ablation_combined.round(5)
            per_evaluator_ablation_combined['dataset'] = 'Combined'
            per_evaluator_ablation_combined = per_evaluator_ablation_combined.reset_index().set_index(['evaluator', 'dataset', 'Ablation'])

            total_stats_ablation_from_master = non_null_ablation_df.groupby(['dataset', 'Ablation']).agg({
                'Accuracy': ['mean', 'std', ('2', lambda x: x.eq(2).sum()), ('1', lambda x: x.eq(1).sum()), ('0', lambda x: x.eq(0).sum())],
                'Helpfulness': ['mean', 'std', ('2', lambda x: x.eq(2).sum()), ('1', lambda x: x.eq(1).sum()), ('0', lambda x: x.eq(0).sum())]
            }).fillna(0)
            if num_evaluators_ablation > 0:
                for col in ['Accuracy', 'Helpfulness']:
                    for score in ['2', '1', '0']:
                        total_stats_ablation_from_master[(col, score)] = total_stats_ablation_from_master[(col, score)] / num_evaluators_ablation
            total_stats_ablation_from_master = total_stats_ablation_from_master.round(5)

            per_evaluator_ablation_counts = stats_ablation[[('Accuracy', 'count'), ('Helpfulness', 'count')]]
            total_ablation_counts = per_evaluator_ablation_counts.groupby(['dataset', 'Ablation']).first()

            total_stats_ablation = pd.merge(total_ablation_counts, total_stats_ablation_from_master, on=['dataset', 'Ablation'], how='outer')
            total_stats_ablation = total_stats_ablation[[
                ('Accuracy', 'count'), ('Accuracy', 'mean'), ('Accuracy', 'std'), ('Accuracy', '2'), ('Accuracy', '1'), ('Accuracy', '0'),
                ('Helpfulness', 'count'), ('Helpfulness', 'mean'), ('Helpfulness', 'std'), ('Helpfulness', '2'), ('Helpfulness', '1'), ('Helpfulness', '0')
            ]]
            total_stats_ablation['evaluator'] = 'Total'
            total_stats_ablation = total_stats_ablation.reset_index().set_index(['evaluator', 'dataset', 'Ablation'])

            overall_total_ablation_from_master = non_null_ablation_df.groupby('Ablation').agg({
                'Accuracy': ['mean', 'std', ('2', lambda x: x.eq(2).sum()), ('1', lambda x: x.eq(1).sum()), ('0', lambda x: x.eq(0).sum())],
                'Helpfulness': ['mean', 'std', ('2', lambda x: x.eq(2).sum()), ('1', lambda x: x.eq(1).sum()), ('0', lambda x: x.eq(0).sum())]
            }).fillna(0)
            if num_evaluators_ablation > 0:
                for col in ['Accuracy', 'Helpfulness']:
                    for score in ['2', '1', '0']:
                        overall_total_ablation_from_master[(col, score)] = overall_total_ablation_from_master[(col, score)] / num_evaluators_ablation
            overall_total_ablation_from_master = overall_total_ablation_from_master.round(5)

            if not total_stats_ablation.empty:
                overall_ablation_counts = total_stats_ablation.groupby('Ablation')[[('Accuracy', 'count'), ('Helpfulness', 'count')]].sum()
            else:
                overall_ablation_counts = per_evaluator_ablation_counts.groupby('Ablation').first()

            overall_total_ablation = pd.merge(overall_ablation_counts, overall_total_ablation_from_master, on='Ablation', how='outer')
            overall_total_ablation = overall_total_ablation[[
                ('Accuracy', 'count'), ('Accuracy', 'mean'), ('Accuracy', 'std'), ('Accuracy', '2'), ('Accuracy', '1'), ('Accuracy', '0'),
                ('Helpfulness', 'count'), ('Helpfulness', 'mean'), ('Helpfulness', 'std'), ('Helpfulness', '2'), ('Helpfulness', '1'), ('Helpfulness', '0')
            ]]
            overall_total_ablation['evaluator'] = 'Total'
            overall_total_ablation['dataset'] = 'Combined'
            overall_total_ablation = overall_total_ablation.reset_index().set_index(['evaluator', 'dataset', 'Ablation'])

            stats_ablation = pd.concat([stats_ablation, per_evaluator_ablation_combined, total_stats_ablation, overall_total_ablation])
            stats_ablation.sort_index(inplace=True)
        else:
            stats_ablation = pd.DataFrame()

    # --- 3. 將所有分析結果寫入一個 Excel 檔案 ---
    print(f"\n正在將分析結果儲存至: {output_filepath}")
    with pd.ExcelWriter(output_filepath, engine='openpyxl') as writer:
        if not stats_method.empty:
            stats_method.to_excel(writer, sheet_name='Method_Statistics')
        if not null_counts_table.empty:
            null_counts_table.to_excel(writer, sheet_name='Ablation_Null_Counts')
        if not stats_ablation.empty:
            stats_ablation.to_excel(writer, sheet_name='Ablation_Non-Null_Stats')

    print("✅ 分析報告儲存成功！")


def perform_additional_ablation_analysis(all_ablation_data, output_filepath):
    """
    執行額外的 Ablation 項目分析 (Preference 和 Crucial Info)，並將結果儲存到新的 Excel 分頁。
    [**新功能**] 在 'Total' 列中，Crucial Info 的計數改為顯示「平均數量」。
    """
    print("\n--- 開始執行額外的 Ablation 分析 ---")
    if not all_ablation_data:
        print("沒有可供分析的 Ablation 資料。")
        return

    master_df = pd.concat(all_ablation_data, ignore_index=True)

    preference_counts_dfs = []

    len_df = master_df[master_df['Ablation'] == 'length = 80'].copy()
    if not len_df.empty:
        len_map = {'50': 'length = 50', 'A': 'length = 50', '80': 'length = 80', 'B': 'length = 80', 'C': 'tie'}
        len_df['Preference_Category'] = len_df['Preference'].astype(str).map(len_map)
        len_counts = len_df.groupby('evaluator')['Preference_Category'].value_counts().unstack(fill_value=0)
        len_counts['Comparison'] = 'length = 50 vs. length = 80'
        preference_counts_dfs.append(len_counts)

    subseq_df = master_df[(master_df['Ablation'] == '+subsequent') & (master_df['dataset'] == 'fce')].copy()
    if not subseq_df.empty:
        subseq_map = {
            '50': 'w/o subsequent (length = 50)',
            'A': 'w/o subsequent (length = 50)',
            'sub': 'w/ subsequent (+subsequent)',
            'subseq': 'w/ subsequent (+subsequent)',
            'B': 'w/ subsequent (+subsequent)',
            'C': 'tie'
        }
        subseq_df['Preference_Category'] = subseq_df['Preference'].astype(str).map(subseq_map)
        subseq_counts = subseq_df.groupby('evaluator')['Preference_Category'].value_counts().unstack(fill_value=0)
        subseq_counts['Comparison'] = 'length = 50 vs. +subsequent'
        preference_counts_dfs.append(subseq_counts)

    final_preference_counts = pd.DataFrame()
    if preference_counts_dfs:
        final_preference_counts = pd.concat(preference_counts_dfs, sort=False).fillna(0)
        final_preference_counts = final_preference_counts.reset_index().set_index(['Comparison', 'evaluator'])

        if not final_preference_counts.empty:
            numeric_cols = final_preference_counts.select_dtypes(include=np.number).columns
            final_preference_counts['Total'] = final_preference_counts[numeric_cols].sum(axis=1)

            avg_pref_counts = final_preference_counts.groupby('Comparison').mean()
            avg_pref_counts['evaluator'] = 'Total'
            avg_pref_counts = avg_pref_counts.reset_index().set_index(['Comparison', 'evaluator'])

            final_preference_counts = pd.concat([final_preference_counts, avg_pref_counts]).sort_index()
            final_preference_counts = final_preference_counts.round(5)

    info_df = master_df.copy()
    info_df['Crucial Information Preservation'] = pd.to_numeric(info_df['Crucial Information Preservation'], errors='coerce')
    info_df.dropna(subset=['Crucial Information Preservation'], inplace=True)

    stats_crucial_info = pd.DataFrame()
    if not info_df.empty:
        num_evaluators_info = info_df['evaluator'].nunique()

        stats_info = info_df.groupby(['evaluator', 'dataset', 'Ablation']).agg({
            'Crucial Information Preservation': ['count', 'mean', 'std', ('2', lambda x: x.eq(2).sum()), ('1', lambda x: x.eq(1).sum()), ('0', lambda x: x.eq(0).sum())]
        }).fillna(0)
        for score in ['2', '1', '0']:
             stats_info[('Crucial Information Preservation', score)] = stats_info[('Crucial Information Preservation', score)].astype(int)
        stats_info = stats_info.round(5)

        per_evaluator_info_combined = info_df.groupby(['evaluator', 'Ablation']).agg({
            'Crucial Information Preservation': ['count', 'mean', 'std', ('2', lambda x: x.eq(2).sum()), ('1', lambda x: x.eq(1).sum()), ('0', lambda x: x.eq(0).sum())]
        }).fillna(0)
        for score in ['2', '1', '0']:
            per_evaluator_info_combined[('Crucial Information Preservation', score)] = per_evaluator_info_combined[('Crucial Information Preservation', score)].astype(int)
        per_evaluator_info_combined = per_evaluator_info_combined.round(5)
        per_evaluator_info_combined['dataset'] = 'Combined'
        per_evaluator_info_combined = per_evaluator_info_combined.reset_index().set_index(['evaluator', 'dataset', 'Ablation'])

        total_stats_info_from_master = info_df.groupby(['dataset', 'Ablation']).agg({
            'Crucial Information Preservation': ['mean', 'std', ('2', lambda x: x.eq(2).sum()), ('1', lambda x: x.eq(1).sum()), ('0', lambda x: x.eq(0).sum())]
        }).fillna(0)
        if num_evaluators_info > 0:
            for score in ['2', '1', '0']:
                total_stats_info_from_master[('Crucial Information Preservation', score)] = total_stats_info_from_master[('Crucial Information Preservation', score)] / num_evaluators_info
        total_stats_info_from_master = total_stats_info_from_master.round(5)

        per_evaluator_counts_info = stats_info[[('Crucial Information Preservation', 'count')]]
        total_counts_info = per_evaluator_counts_info.groupby(['dataset', 'Ablation']).first()
        total_stats_info = pd.merge(total_counts_info, total_stats_info_from_master, on=['dataset', 'Ablation'], how='outer')
        total_stats_info['evaluator'] = 'Total'
        total_stats_info = total_stats_info.reset_index().set_index(['evaluator', 'dataset', 'Ablation'])

        overall_total_stats_info_master = info_df.groupby('Ablation').agg({
            'Crucial Information Preservation': ['mean', 'std', ('2', lambda x: x.eq(2).sum()), ('1', lambda x: x.eq(1).sum()), ('0', lambda x: x.eq(0).sum())]
        }).fillna(0)
        if num_evaluators_info > 0:
            for score in ['2', '1', '0']:
                overall_total_stats_info_master[('Crucial Information Preservation', score)] = overall_total_stats_info_master[('Crucial Information Preservation', score)] / num_evaluators_info
        overall_total_stats_info_master = overall_total_stats_info_master.round(5)

        if not total_stats_info.empty:
            overall_total_counts_info = total_stats_info.groupby('Ablation')[[('Crucial Information Preservation', 'count')]].sum()
        else:
            overall_total_counts_info = per_evaluator_counts_info.groupby('Ablation').first()

        overall_total_info = pd.merge(overall_total_counts_info, overall_total_stats_info_master, on='Ablation', how='outer')
        overall_total_info['evaluator'] = 'Total'
        overall_total_info['dataset'] = 'Combined'
        overall_total_info = overall_total_info.reset_index().set_index(['evaluator', 'dataset', 'Ablation'])

        stats_crucial_info = pd.concat([stats_info, per_evaluator_info_combined, total_stats_info, overall_total_info])
        stats_crucial_info.sort_index(inplace=True)

    try:
        with pd.ExcelWriter(output_filepath, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
            if not final_preference_counts.empty:
                final_preference_counts.to_excel(writer, sheet_name='Ablation_Preference_Counts')
            if not stats_crucial_info.empty:
                stats_crucial_info.to_excel(writer, sheet_name='Ablation_Info_Preservation_Stats')
        print(f"額外的 Ablation 分析已成功儲存至 '{output_filepath}'。")
    except FileNotFoundError:
        print(f"錯誤：找不到目標檔案 '{output_filepath}'。請先執行 generate_summary_report()。")
    except Exception as e:
        print(f"儲存額外的 Ablation 分析時發生錯誤：{e}")


def analyze_by_sentence_category(all_method_data, output_filepath):
    """
    [**已重構**] 根據提供的句子分類，統計評估結果，並將結果存到新的 Excel 分頁。
    - **重新加入 Total 統計，並擴充統計指標 (count, mean, std, 2, 1, 0)。**
    - 'count' 代表不重複的句子數量。
    - 會驗證 FCE 與 Longman 是否各有 20 句不重複的句子。
    - 四捨五入至小數點後第五位。
    """
    print("\n--- 開始依據句子類型進行分析 ---")
    if not all_method_data:
        print("沒有可供分析的『方法』資料。")
        return

    fce_types = {
        'RV': ["This made me {+feel+}[-get-] bored. + You also say in your advertisement that there are some discounts available but there weren't any.",
               "After that Paul {+turned+}[-got-] white in the face, he was pale like a wall and it seemed that he was  very nervous. + Pat couldn't understand Paul's reaction and so she didn't worry about it.",
               'One of our teachers, who will be {+elected+}[-voted-] by students will make a speech for one hour. + After, a student will make a speech about what he or she has learned, and dinner.',
               'AS SOME OF THEM ARE CONSIDERED IDOLS, THEIR FANS WANT TO KNOW HOW THEY {+BEHAVE+}[-ACT-] IN THEIR PRIVATE LIVES, WITH THEIR FAMILY OR FRIENDS. + AS A RESPONSE TO THIS DEMAND FOR INFORMATION, JOURNALISTS FOLLOW FAMOUS PEOPLE DAY AND NIGHT IN MANY DIFFERENT WAYS, FOR EXAMPLE, CHASING THEIR CARS, TAKING PHOTOGRAPHS OR BREAKING INTO THEIR HOUSES WHILE THEY ARE ON HOLIDAY.',
               'They become lazy because they know that they can {+sit+}[-seat-] on the sofa and change the channels on the TV by pressing a button. + In conclusion, I want to say that in spite of this, I hope that in future we will have more and more new inventions and they will prevent our disasters and problems.',
               'Modern technology is {+affecting+}[-effecting-] people in a good and bad way. + We have to be very careful with technology.', 'First of all, the reason that we decided to {+have+}[-do-] this party was because Mr Brown helped with the organisation of our trip to France. + And we would appreciate your not telling him about our party.',
               'Could you tell me  the amount of money I should {+have+}[-bring-] during the holiday? + I look forward to hearing from you!', 'As for singing, I am very keen on singing as I {+take+}[-get-] lessons in it. + I just love singing.'],
        'RJ': ['But there are some drawbacks, which appeared in {+recent+}[-last-] years. + For example, a lot of pollution has been created.',
               'AS REGARDS THE TICKETS, I READ THAT DISCOUNTS WERE AVAILABLE, BUT IT WAS NOT SO, AND I HAD TO PAY THE {+FULL+}[-WHOLE-] PRICE. + MOREOVER, AFTER THE SHOW, I INTENDED TO VISIT YOUR THEATRE RESTAURANT, BUT IT WAS CLOSED AND THERE WAS NO EXPLANATION FOR THAT.',
               "You can't imagine how {+fun+}[-funny-] it was! + You could do so many things because you received a free pass which allowed you to go where you wanted and that was what I liked about this experience.",
               "Regarding the choice of tents or log cabins, I think I will turn to the first one because I have always loved camping {+close+}[-the nearest-] to nature, if  possible, in front of a lake or a river, if your campsite has one, so that I will be able to do my favourite and skillful hobby: sailing. + As well as this activity I'll do painting, I'm very keen on art and I think I'm good at painting portraits.",
               'On your last day you could go to the sports museum, which is {+great+}[-grateful-] and famous around the world, or just go for a walk around Reconvillier Park. + I hope that I have answered all of your questions and that the conference will be great.'],
        'RN': ['The problem with this job is that you have to deal with huge {+numbers+}[-amounts-] of people. + There were 963 people at the "Armageddon" that night!', 'Secondly, a {+camping+}[-tent-] holiday has always attracted me so I would like to stay in a tent rather than in a log cabin. + In my spare time, I enjoy playing basketball with my friends and I also like swimming.',
               'To conclude, although you can have good {+times+}[-moments-] while shopping all of us know that it is not always enjoyable, is it?',
               'This trip would give us the {+chance+}[-change-] to see London and to improve our English . + We all like the daily programme for the three days.'],
        'L': ["I can't explain how {+great+}[-super-] it was. + I stayed with him until the concert began and after the show I had the opportunity to stay in his private room with him and the other dancers."],
        'RY': ['I would {+instead+}[-rather-] say that it was a very disappointing evening, and I would like to have my money back as soon as possible or we will meet in court! + Yours sincerely,']
    }

    longman_types = {
        'one_error': {
            'VERB': ['Mr Tong was {+appointed+}[-nominated-] manager of the company in 1984.',
                     'The story {+revolves+}[-circulates-] around his career in the army.',
                     'This morning I noticed that my purse was {+missing+}[-lost-].',
                     'Kiri was always {+talking+}[-telling-] about herself and her problems.',
                     'The car was {+moving+}[-running-] too fast for me to see the number plate.',
                     'The public are {+demanding+}[-claiming-] stricter laws.',
                     'The headmaster {+carried out+}[-realized-] his threat and sent the children home.'],
            'NOUN': ['The wheat is collected and sent to the flour {+mill+}[-factory-].',
                     'Her husband talked so much that all the other {+people+}[-persons-] in the room had to keep quiet.'],
            'ADJ': ['If I did the same thing every day, I would be {+bored+}[-dull-].',
                    'The food was excellent and very {+tasty+}[-tasteful-].', 'The police asked for a {+brief+}[-little-] description of the car.'],
            'ADV': ['The poor bus service makes it difficult to get {+anywhere+}[-somewhere-].',
                    "I'll phone you {+back+}[-again-] in five minutes."],
            'OTHER': ['I jumped in and swam towards the child {+as fast as I could+}[-at my fastest speed-].',
                      'She was standing by the reception desk, {+waiting for+}[-expecting-] a taxi.',
                      'New doors cost {+a lot of money+}[-very much-] because wood is so expensive.',
                      "It's {+raining+}[-rainy-] again today."]
        },
        'multi_error': {
            'R': ['We {+saw+}[-visited-] all the famous {+sights+}[-sightseeing places-].'],
            'M': [],
            'U': [],
            'Mix': ['This decision changed [-all of-] her {+whole+} life.']
        }
    }

    fce_map = {s: cat for cat, sentences in fce_types.items() for s in sentences}
    longman_map = {s: (main_cat, sub_cat) for main_cat, sub_dict in longman_types.items() for sub_cat, sentences in sub_dict.items() for s in sentences}

    master_df = pd.concat(all_method_data, ignore_index=True)

    def find_category_details(row):
        key_sentence = row['key_sentence']
        dataset = row['dataset']
        if pd.isna(key_sentence):
            return None, None
        if dataset == 'fce':
            cat = fce_map.get(key_sentence, None)
            return 'FCE', cat
        elif dataset == 'longman':
            return longman_map.get(key_sentence, (None, None))
        return None, None

    master_df[['Main_Category', 'Sub_Category']] = master_df.apply(find_category_details, axis=1, result_type='expand')
    master_df['Accuracy'] = pd.to_numeric(master_df['Accuracy'], errors='coerce')
    master_df['Helpfulness'] = pd.to_numeric(master_df['Helpfulness'], errors='coerce')

    # --- 共通的統計設定 ---
    agg_metrics = {
        'Accuracy': ['mean', 'std', ('2', lambda x: x.eq(2).sum()), ('1', lambda x: x.eq(1).sum()), ('0', lambda x: x.eq(0).sum())],
        'Helpfulness': ['mean', 'std', ('2', lambda x: x.eq(2).sum()), ('1', lambda x: x.eq(1).sum()), ('0', lambda x: x.eq(0).sum())]
    }
    cols_order = [
        ('Accuracy', 'count'), ('Accuracy', 'mean'), ('Accuracy', 'std'), ('Accuracy', '2'), ('Accuracy', '1'), ('Accuracy', '0'),
        ('Helpfulness', 'count'), ('Helpfulness', 'mean'), ('Helpfulness', 'std'), ('Helpfulness', '2'), ('Helpfulness', '1'), ('Helpfulness', '0')
    ]

    # --- 分析 FCE 資料 ---
    fce_df = master_df[(master_df['dataset'] == 'fce') & (master_df['Sub_Category'].notna())].copy()
    stats_fce = pd.DataFrame()
    if not fce_df.empty:
        fce_sentence_count = fce_df['key_sentence'].nunique()
        print(f"  - 正在分析 {fce_sentence_count} 句不重複的 FCE 句子 (預期為 20 句)。")

        # 計算各類別的統計數據
        stats_fce = fce_df.groupby(['Sub_Category', 'Method']).agg(agg_metrics).fillna(0)
        counts_fce = fce_df.groupby(['Sub_Category', 'Method'])['key_sentence'].nunique()
        stats_fce[('Accuracy', 'count')] = counts_fce
        stats_fce[('Helpfulness', 'count')] = counts_fce
        stats_fce = stats_fce[cols_order].round(5)

        # 計算 FCE 總計
        fce_total_stats = fce_df.groupby('Method').agg(agg_metrics).fillna(0)
        fce_total_stats[('Accuracy', 'count')] = fce_sentence_count
        fce_total_stats[('Helpfulness', 'count')] = fce_sentence_count
        fce_total_stats['Sub_Category'] = 'Total (FCE)'
        fce_total = fce_total_stats.reset_index().set_index(['Sub_Category', 'Method'])
        fce_total = fce_total[cols_order].round(5)

        # 合併結果
        stats_fce = pd.concat([stats_fce, fce_total])
        stats_fce.index.names = ['Category', 'Method']

    # --- 分析 Longman 資料 ---
    longman_df = master_df[(master_df['dataset'] == 'longman') & (master_df['Sub_Category'].notna())].copy()
    stats_longman = pd.DataFrame()
    if not longman_df.empty:
        longman_sentence_count = longman_df['key_sentence'].nunique()
        print(f"  - 正在分析 {longman_sentence_count} 句不重複的 Longman 句子 (預期為 20 句)。")

        # 計算各類別的統計數據
        stats_longman = longman_df.groupby(['Main_Category', 'Sub_Category', 'Method']).agg(agg_metrics).fillna(0)
        counts_longman = longman_df.groupby(['Main_Category', 'Sub_Category', 'Method'])['key_sentence'].nunique()
        stats_longman[('Accuracy', 'count')] = counts_longman
        stats_longman[('Helpfulness', 'count')] = counts_longman
        stats_longman = stats_longman[cols_order].round(5)

        # 計算 Longman 總計
        longman_total_stats = longman_df.groupby('Method').agg(agg_metrics).fillna(0)
        longman_total_stats[('Accuracy', 'count')] = longman_sentence_count
        longman_total_stats[('Helpfulness', 'count')] = longman_sentence_count
        longman_total_stats['Main_Category'] = 'Total'
        longman_total_stats['Sub_Category'] = 'Total (Longman)'
        longman_total = longman_total_stats.reset_index().set_index(['Main_Category', 'Sub_Category', 'Method'])
        longman_total = longman_total[cols_order].round(5)

        # 合併結果
        stats_longman = pd.concat([stats_longman, longman_total])

    # --- 將結果寫入 Excel ---
    try:
        with pd.ExcelWriter(output_filepath, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
            if not stats_fce.empty:
                stats_fce.to_excel(writer, sheet_name='Analysis_by_FCE_Type')
            if not stats_longman.empty:
                stats_longman.to_excel(writer, sheet_name='Analysis_by_Longman_Type')
        print(f"句子類型分析已成功儲存至 '{output_filepath}'。")
    except FileNotFoundError:
        print(f"錯誤：找不到目標檔案 '{output_filepath}'。請先執行 generate_summary_report()。")
    except Exception as e:
        print(f"儲存句子類型分析時發生錯誤：{e}")


def perform_t_tests(all_method_data, output_filepath):
    """
    對不同系統的評估結果執行成對 t-test，並將結果儲存到新的 Excel 分頁。
    [**修改**] 四捨五入至小數點後第五位。
    """
    print("\n--- 開始執行成對 T-test ---")
    if not all_method_data:
        print("沒有可用於 T-test 的資料。")
        return

    master_df = pd.concat(all_method_data, ignore_index=True).copy()
    master_df['Accuracy'] = pd.to_numeric(master_df['Accuracy'], errors='coerce')
    master_df['Helpfulness'] = pd.to_numeric(master_df['Helpfulness'], errors='coerce')

    all_results = []

    def run_and_format_ttest(df, dataset_name, metric):
        pivot_df = df.pivot_table(index=['key_sentence', 'evaluator'], columns='Method', values=metric, aggfunc='first').dropna()

        systems_to_test = [col for col in pivot_df.columns if col not in ['original', 'baseline']]

        comparisons = []
        if 'original' in pivot_df.columns:
            if dataset_name == 'Longman':
                comparisons.extend([('original', system) for system in systems_to_test])

        if 'baseline' in pivot_df.columns:
            if dataset_name == 'FCE' or dataset_name == 'Total (Combined)':
                comparisons.extend([('baseline', system) for system in systems_to_test])

        for base_system, compare_system in comparisons:
            if base_system in pivot_df.columns and compare_system in pivot_df.columns:
                stat, p_value = ttest_rel(pivot_df[base_system], pivot_df[compare_system])
                all_results.append({
                    'Dataset': dataset_name,
                    'Metric': metric,
                    'Comparison': f"{base_system} vs {compare_system}",
                    'T-Statistic': round(stat, 5),
                    'P-Value': round(p_value, 5)
                })

    longman_df = master_df[master_df['dataset'] == 'longman']
    if not longman_df.empty:
        run_and_format_ttest(longman_df, 'Longman', 'Accuracy')
        run_and_format_ttest(longman_df, 'Longman', 'Helpfulness')

    fce_df = master_df[master_df['dataset'] == 'fce']
    if not fce_df.empty:
        run_and_format_ttest(fce_df, 'FCE', 'Accuracy')
        run_and_format_ttest(fce_df, 'FCE', 'Helpfulness')

    if not master_df.empty:
        run_and_format_ttest(master_df, 'Total (Combined)', 'Accuracy')
        run_and_format_ttest(master_df, 'Total (Combined)', 'Helpfulness')

    if not all_results:
        print("沒有足夠的成對資料來執行 T-test。")
        return

    results_df = pd.DataFrame(all_results)

    try:
        with pd.ExcelWriter(output_filepath, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
            results_df.to_excel(writer, sheet_name='T-Test_Results', index=False)
        print(f"T-test 分析已成功儲存至 '{output_filepath}' 的 'T-Test_Results' 分頁。")
    except FileNotFoundError:
        print(f"錯誤：找不到目標檔案 '{output_filepath}'。請先執行 generate_summary_report()。")
    except Exception as e:
        print(f"儲存 T-test 分析時發生錯誤：{e}")


def perform_ablation_t_tests(all_method_data, all_ablation_data, output_filepath):
    """
    [**已重構**] 對 Ablation 實驗組與 'dictionary' 基準組進行成對 T-test。
    - **新增對 `Crucial Information Preservation` 的 T-test。**
    - 四捨五入至小數點後第五位。
    """
    print("\n--- 開始執行 Ablation T-test ---")
    if not all_ablation_data or not all_method_data:
        print("沒有足夠的資料來執行 Ablation T-test。")
        return

    master_method_df = pd.concat(all_method_data, ignore_index=True)
    master_ablation_df = pd.concat(all_ablation_data, ignore_index=True)

    baseline_df = master_method_df[master_method_df['Method'] == 'dictionary'].copy()
    baseline_df['Accuracy'] = pd.to_numeric(baseline_df['Accuracy'], errors='coerce')
    baseline_df['Helpfulness'] = pd.to_numeric(baseline_df['Helpfulness'], errors='coerce')

    master_ablation_df['Accuracy'] = pd.to_numeric(master_ablation_df['Accuracy'], errors='coerce')
    master_ablation_df['Helpfulness'] = pd.to_numeric(master_ablation_df['Helpfulness'], errors='coerce')
    master_ablation_df['Crucial Information Preservation'] = pd.to_numeric(master_ablation_df['Crucial Information Preservation'], errors='coerce')

    all_results = []

    def run_and_format_ablation_ttest(comparison_df, baseline_df, dataset_filter, dataset_label, metric, comparison_label, results_list):
        if metric == 'Crucial Information Preservation':
            print(f"    - 資訊: '{metric}' 僅存在於 Ablation 資料中，無法與 Method 資料中的 'dictionary' 進行 T-test 比較，已跳過。")
            return

        if dataset_filter != 'all':
            comp_data = comparison_df[comparison_df['dataset'] == dataset_filter]
            base_data = baseline_df[baseline_df['dataset'] == dataset_filter]
        else:
            comp_data = comparison_df
            base_data = baseline_df

        merged_df = pd.merge(
            base_data[['key_sentence', 'evaluator', metric]],
            comp_data[['key_sentence', 'evaluator', metric]],
            on=['key_sentence', 'evaluator'],
            suffixes=('_base', '_comp')
        ).dropna()

        if len(merged_df) < 2:
            print(f"    - 警告: 對於 {dataset_label} - {comparison_label} - {metric}，找不到足夠的成對資料 (僅 {len(merged_df)} 對)，已跳過。")
            return

        stat, p_value = ttest_rel(merged_df[f'{metric}_base'], merged_df[f'{metric}_comp'])
        results_list.append({
            'Dataset': dataset_label,
            'Metric': metric,
            'Comparison': comparison_label,
            'T-Statistic': round(stat, 5),
            'P-Value': round(p_value, 5),
            'Paired_Samples': len(merged_df)
        })

    comparisons_to_run = {
        'dictionary vs length = 80': master_ablation_df[master_ablation_df['Ablation'] == 'length = 80'],
        'dictionary vs +subsequent': master_ablation_df[master_ablation_df['Ablation'] == '+subsequent']
    }

    metrics_to_test = ['Accuracy', 'Helpfulness', 'Crucial Information Preservation']

    for comp_label, comp_df in comparisons_to_run.items():
        for metric in metrics_to_test:
            run_and_format_ablation_ttest(comp_df, baseline_df, 'fce', 'FCE', metric, comp_label, all_results)
            run_and_format_ablation_ttest(comp_df, baseline_df, 'longman', 'Longman', metric, comp_label, all_results)
            run_and_format_ablation_ttest(comp_df, baseline_df, 'all', 'Combined', metric, comp_label, all_results)

    if not all_results:
        print("Ablation T-test 未產生任何結果。")
        return

    results_df = pd.DataFrame(all_results)

    try:
        with pd.ExcelWriter(output_filepath, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
            results_df.to_excel(writer, sheet_name='Ablation_T-Test_Results', index=False)
        print(f"Ablation T-test 分析已成功儲存至 '{output_filepath}' 的 'Ablation_T-Test_Results' 分頁。")
    except FileNotFoundError:
        print(f"錯誤：找不到目標檔案 '{output_filepath}'。請先執行 generate_summary_report()。")
    except Exception as e:
        print(f"儲存 Ablation T-test 分析時發生錯誤：{e}")

def perform_preference_chi_squared_test(all_ablation_data, output_filepath):
    """
    [**新功能**] 對 Preference 欄位的計數進行卡方檢定 (Chi-squared Test)。
    此函式用於檢定不同偏好選項的次數差異是否具有統計顯著性。
    """
    print("\n--- 開始執行 Preference 卡方檢定 ---")
    if not all_ablation_data:
        print("沒有可分析的 Ablation 資料。")
        return

    master_df = pd.concat(all_ablation_data, ignore_index=True)
    results = []

    # --- 比較 1: length = 50 vs. length = 80 ---
    len_df = master_df[master_df['Ablation'] == 'length = 80'].copy()
    if not len_df.empty:
        len_map = {'50': 'length = 50', 'A': 'length = 50', '80': 'length = 80', 'B': 'length = 80', 'C': 'tie'}
        len_df['Preference_Category'] = len_df['Preference'].astype(str).map(len_map)
        counts = len_df['Preference_Category'].value_counts()

        count1 = counts.get('length = 50', 0)
        count2 = counts.get('length = 80', 0)

        if count1 + count2 > 0:
            stat, p_value = chisquare([count1, count2])
            results.append({
                "Comparison": "length = 50 vs. length = 80",
                "Preference_1": "length = 50",
                "Count_1": count1,
                "Preference_2": "length = 80",
                "Count_2": count2,
                "Ties": counts.get('tie', 0),
                "Chi-Squared_Statistic": round(stat, 5),
                "P-Value": round(p_value, 5)
            })
        else:
            print("  - 在 'length' 比較中未找到足夠的偏好計數。")

    # --- 比較 2: w/o subsequent vs. +subsequent (僅 FCE) ---
    subseq_df = master_df[(master_df['Ablation'] == '+subsequent') & (master_df['dataset'] == 'fce')].copy()
    if not subseq_df.empty:
        subseq_map = {
            '50': 'w/o subsequent', 'A': 'w/o subsequent',
            'sub': 'w/ subsequent', 'subseq': 'w/ subsequent', 'B': 'w/ subsequent',
            'C': 'tie'
        }
        subseq_df['Preference_Category'] = subseq_df['Preference'].astype(str).map(subseq_map)
        counts = subseq_df['Preference_Category'].value_counts()

        count1 = counts.get('w/o subsequent', 0)
        count2 = counts.get('w/ subsequent', 0)

        if count1 + count2 > 0:
            stat, p_value = chisquare([count1, count2])
            results.append({
                "Comparison": "w/o subsequent vs. w/ subsequent (FCE)",
                "Preference_1": "w/o subsequent",
                "Count_1": count1,
                "Preference_2": "w/ subsequent",
                "Count_2": count2,
                "Ties": counts.get('tie', 0),
                "Chi-Squared_Statistic": round(stat, 5),
                "P-Value": round(p_value, 5)
            })
        else:
            print("  - 在 'subsequent' 比較中未找到足夠的偏好計數。")

    if not results:
        print("未產生任何 Preference 檢定結果。")
        return

    results_df = pd.DataFrame(results)
    try:
        with pd.ExcelWriter(output_filepath, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
            results_df.to_excel(writer, sheet_name='Ablation_Preference_Test', index=False)
        print(f"Preference 卡方檢定分析已成功儲存至 '{output_filepath}'。")
    except FileNotFoundError:
        print(f"錯誤：找不到目標檔案 '{output_filepath}'。請先執行 generate_summary_report()。")
    except Exception as e:
        print(f"儲存 Preference 卡方檢定分析時發生錯誤：{e}")


def calculate_irr_metrics(scores1, scores2, metric_name):
    """
    [**已重構**] 計算多個評分者信度指標，可處理數值和類別型資料。
    """
    # 判斷是否為有序的數值型指標
    is_numeric_metric = metric_name in ['Accuracy', 'Helpfulness', 'Crucial Information Preservation']

    if is_numeric_metric:
        s1 = np.asarray(scores1, dtype=int)
        s2 = np.asarray(scores2, dtype=int)
    else: # 處理類別型指標 (如 Preference)
        s1 = np.asarray(scores1, dtype=str)
        s2 = np.asarray(scores2, dtype=str)

    # 計算一致性百分比
    agreement_count = np.sum(s1 == s2)
    total_count = len(s1)
    percent_agreement = agreement_count / total_count if total_count > 0 else 0

    # 取得所有評分類別
    labels = sorted(list(set(np.concatenate((s1, s2)))))
    if not labels or len(labels) < 2:
        return {
            "Percent Agreement": round(percent_agreement, 5) if total_count > 0 else 0,
            "Cohen's Kappa": np.nan,
            "Weighted Kappa (Quadratic)": np.nan,
            "Gwet's AC1": np.nan
        }

    # 計算 Cohen's Kappa (無加權)
    unweighted_kappa = cohen_kappa_score(s1, s2, labels=labels)

    # 僅對數值型指標計算加權 Kappa
    if is_numeric_metric:
        weighted_kappa = cohen_kappa_score(s1, s2, labels=labels, weights='quadratic')
    else:
        weighted_kappa = np.nan # 對類別型資料無意義

    # 計算 Gwet's AC1
    all_ratings = np.concatenate((s1, s2))
    category_proportions = {cat: np.sum(all_ratings == cat) / len(all_ratings) for cat in labels}
    pe_gwet = sum(p**2 for p in category_proportions.values())
    ac1_denominator = 1 - pe_gwet
    gwet_ac1 = (percent_agreement - pe_gwet) / ac1_denominator if ac1_denominator > 0 else 0

    return {
        "Percent Agreement": round(percent_agreement, 5),
        "Cohen's Kappa": round(unweighted_kappa, 5),
        "Weighted Kappa (Quadratic)": np.nan if np.isnan(weighted_kappa) else round(weighted_kappa, 5),
        "Gwet's AC1": round(gwet_ac1, 5)
    }


def perform_method_irr_analysis(all_method_data, output_filepath):
    """
    [**已重構**] 計算 Method 資料的多個評分者信度 (IRR) 指標。
    """
    print("\n--- 開始執行 Method 評分者信度分析 (IRR) ---")
    if not all_method_data:
        print("沒有可用於 Method IRR 分析的資料。")
        return

    master_df = pd.concat(all_method_data, ignore_index=True).copy()
    master_df['Accuracy'] = pd.to_numeric(master_df['Accuracy'], errors='coerce')
    master_df['Helpfulness'] = pd.to_numeric(master_df['Helpfulness'], errors='coerce')

    all_results = []

    for metric in ['Accuracy', 'Helpfulness']:
        metric_df = master_df.dropna(subset=[metric])
        pivot_df = metric_df.pivot_table(index=['key_sentence', 'Method'], columns='evaluator', values=metric, aggfunc='first')
        evaluators = pivot_df.columns

        if len(evaluators) < 2:
            continue

        # --- 5. 兩兩比較所有評估者 ---
        for rater1, rater2 in combinations(evaluators, 2):
            paired_scores = pivot_df[[rater1, rater2]].dropna()
            if len(paired_scores) > 1:
                scores1 = paired_scores[rater1]
                scores2 = paired_scores[rater2]

                irr_results = calculate_irr_metrics(scores1, scores2, metric)

                result_row = {
                    'Metric': metric,
                    'Evaluator 1': rater1,
                    'Evaluator 2': rater2,
                    'Paired_Ratings': len(paired_scores),
                }
                result_row.update(irr_results)
                all_results.append(result_row)

    if not all_results:
        print("沒有足夠的共同評分項目來計算 Method 的 IRR。")
        return

    results_df = pd.DataFrame(all_results)

    try:
        with pd.ExcelWriter(output_filepath, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
            results_df.to_excel(writer, sheet_name='Method_IRR_Analysis', index=False)
        print(f"Method IRR 分析已成功儲存至 '{output_filepath}' 的 'Method_IRR_Analysis' 分頁。")
    except FileNotFoundError:
        print(f"錯誤：找不到目標檔案 '{output_filepath}'。請先執行 generate_summary_report()。")
    except Exception as e:
        print(f"儲存 Method IRR 分析時發生錯誤：{e}")


def perform_ablation_irr_analysis(all_ablation_data, output_filepath):
    """
    [**已重構**] 計算 Ablation 資料的多個評分者信度 (IRR) 指標。
    參考 generate_summary_report() 與 perform_additional_ablation_analysis() 的指標計算方式，
    採用相同的資料處理邏輯和統計計算模式。
    - 支援對 `Accuracy`, `Helpfulness`, `Crucial Information Preservation` 和 `Preference` 的分析。
    - 針對每個獨立的 Ablation 類型進行 IRR 分析。
    - 遵循相同的數值轉換和缺失值處理邏輯。
    """
    print("\n--- 開始執行 Ablation 評分者信度分析 (IRR) ---")

    if not all_ablation_data:
        print("沒有可供 Ablation IRR 分析的資料。")
        return

    # --- 1. 資料預處理 (參考 generate_summary_report 的處理方式) ---
    master_df = pd.concat(all_ablation_data, ignore_index=True).copy()

    # 處理 Content 欄位的空值 (與 generate_summary_report 一致)
    master_df['Content'] = master_df['Content'].replace(r'^\s*(\[\])?\s*$', np.nan, regex=True)

    # 數值欄位轉換 (與 perform_additional_ablation_analysis 一致)
    master_df['Accuracy'] = pd.to_numeric(master_df['Accuracy'], errors='coerce')
    master_df['Helpfulness'] = pd.to_numeric(master_df['Helpfulness'], errors='coerce')
    master_df['Crucial Information Preservation'] = pd.to_numeric(master_df['Crucial Information Preservation'], errors='coerce')
    master_df['Preference'] = master_df['Preference'].astype(str)

    all_results = []
    metrics_to_analyze = ['Accuracy', 'Helpfulness', 'Crucial Information Preservation', 'Preference']

    # --- 2. 取得所有獨立的 Ablation 類型 ---
    ablation_types = master_df['Ablation'].unique()

    # --- 3. 遍歷每種 Ablation 類型，並獨立計算其 IRR ---
    for ablation_type in ablation_types:
        print(f"  - 正在分析 Ablation 類型: {ablation_type}")
        ablation_df = master_df[master_df['Ablation'] == ablation_type].copy()

        if ablation_df.empty:
            continue

        # --- 4. 遍歷每個指標，計算 IRR ---
        for metric in metrics_to_analyze:
            # 針對不同指標採用適當的資料過濾方式
            if metric in ['Accuracy', 'Helpfulness']:
                # 對於 Accuracy 和 Helpfulness，過濾掉 Content 為空的資料 (與 generate_summary_report 一致)
                metric_df = ablation_df.dropna(subset=['Content', metric]).copy()
            elif metric == 'Crucial Information Preservation':
                # 對於 Crucial Information Preservation，只需過濾該欄位 (與 perform_additional_ablation_analysis 一致)
                metric_df = ablation_df.dropna(subset=[metric]).copy()
            else:  # Preference
                # 對於 Preference，處理方式與其他指標相同
                metric_df = ablation_df.dropna(subset=[metric]).copy()

            if metric_df.empty:
                print(f"    - 警告: 在 '{ablation_type}' 中, '{metric}' 欄位沒有足夠資料，已跳過。")
                continue

            # 將資料轉換為以 key_sentence 為索引，評估者為欄位的格式
            try:
                pivot_df = metric_df.pivot_table(
                    index='key_sentence',
                    columns='evaluator',
                    values=metric,
                    aggfunc='first'  # 處理重複評分的情況
                )
            except Exception as e:
                print(f"    - 錯誤: 為 '{metric}' 建立 pivot table 時失敗: {e}")
                continue

            evaluators = pivot_df.columns
            if len(evaluators) < 2:
                continue

            # --- 5. 兩兩比較所有評估者 (與 perform_method_irr_analysis 一致) ---
            for rater1, rater2 in combinations(evaluators, 2):
                paired_scores = pivot_df[[rater1, rater2]].dropna()

                if len(paired_scores) > 1:
                    scores1 = paired_scores[rater1]
                    scores2 = paired_scores[rater2]

                    # 計算 IRR 指標
                    irr_results = calculate_irr_metrics(scores1, scores2, metric)

                    result_row = {
                        'Ablation_Type': ablation_type,
                        'Metric': metric,
                        'Evaluator 1': rater1,
                        'Evaluator 2': rater2,
                        'Paired_Ratings': len(paired_scores),
                    }
                    result_row.update(irr_results)
                    all_results.append(result_row)

    if not all_results:
        print("沒有足夠的共同評分項目來計算 Ablation 的 IRR。")
        return

    # --- 6. 將結果儲存至 Excel (與其他函數一致的儲存方式) ---
    results_df = pd.DataFrame(all_results)
    try:
        with pd.ExcelWriter(output_filepath, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
            results_df.to_excel(writer, sheet_name='Ablation_IRR_Analysis', index=False)
        print(f"Ablation IRR 分析已成功儲存至 '{output_filepath}' 的 'Ablation_IRR_Analysis' 分頁。")
    except FileNotFoundError:
        print(f"錯誤：找不到目標檔案 '{output_filepath}'。請先執行 generate_summary_report()。")
    except Exception as e:
        print(f"儲存 Ablation IRR 分析時發生錯誤：{e}")


def select_qualitative_analysis_examples(all_method_data, all_ablation_data, output_filepath):
    """
    [**已重構**] 根據評估結果挑選用於質性分析的代表性例子。
    - 包含 Method 和 Ablation 研究的案例。
    - 採用更詳細的系統比較規則。
    - 優化案例選擇邏輯，確保一個案例可以包含多個被選中的理由。
    """
    print("\n--- 開始挑選質性分析案例 ---")

    if not all_method_data and not all_ablation_data:
        print("沒有可用於質性分析的資料。")
        return

    master_method_df = pd.DataFrame()
    if all_method_data:
        master_method_df = pd.concat(all_method_data, ignore_index=True).copy()
        master_method_df['Accuracy'] = pd.to_numeric(master_method_df['Accuracy'], errors='coerce')
        master_method_df['Helpfulness'] = pd.to_numeric(master_method_df['Helpfulness'], errors='coerce')

    master_ablation_df = pd.DataFrame()
    if all_ablation_data:
        master_ablation_df = pd.concat(all_ablation_data, ignore_index=True).copy()
        for col in ['Accuracy', 'Helpfulness', 'Crucial Information Preservation']:
            master_ablation_df[col] = pd.to_numeric(master_ablation_df[col], errors='coerce')

    selected_sentences = {}

    def add_selection(key, reason, selection_dict):
        if key not in selection_dict:
            selection_dict[key] = []
        if reason not in selection_dict[key]:
            selection_dict[key].append(reason)

    if not master_method_df.empty:
        method_stats = master_method_df.groupby(['key_sentence', 'dataset', 'Method']).agg(
            Avg_Accuracy=('Accuracy', 'mean'),
            Std_Accuracy=('Accuracy', 'std'),
        ).reset_index().fillna(0)

        overall_avg = method_stats.groupby('key_sentence')['Avg_Accuracy'].mean()
        if not overall_avg.empty:
            add_selection(overall_avg.idxmax(), "Method: 表現最佳 (平均分數最高)", selected_sentences)
            add_selection(overall_avg.idxmin(), "Method: 表現最差 (平均分數最低)", selected_sentences)

        if not method_stats.empty:
            most_disagreement_key = method_stats.loc[method_stats['Std_Accuracy'].idxmax()]['key_sentence']
            add_selection(most_disagreement_key, "Method: 評估者意見分歧最大", selected_sentences)

        longman_stats = method_stats[method_stats['dataset'] == 'longman']
        if not longman_stats.empty:
            pivot_longman = longman_stats.pivot_table(index='key_sentence', columns='Method', values='Avg_Accuracy')
            systems_vs_original = ['baseline', 'dictionary', 'collocation', 'metalinguistic', 'mix']
            if 'original' in pivot_longman.columns:
                for system in systems_vs_original:
                    if system in pivot_longman.columns:
                        diff = (pivot_longman['original'] - pivot_longman[system]).abs().dropna()
                        if not diff.empty:
                            add_selection(diff.idxmax(), f"系統差異最大 (Longman): original vs {system}", selected_sentences)
            systems_vs_baseline = ['dictionary', 'collocation', 'metalinguistic', 'mix']
            if 'baseline' in pivot_longman.columns:
                for system in systems_vs_baseline:
                    if system in pivot_longman.columns:
                        diff = (pivot_longman['baseline'] - pivot_longman[system]).abs().dropna()
                        if not diff.empty:
                            add_selection(diff.idxmax(), f"系統差異最大 (Longman): baseline vs {system}", selected_sentences)

        fce_stats = method_stats[method_stats['dataset'] == 'fce']
        if not fce_stats.empty:
            pivot_fce = fce_stats.pivot_table(index='key_sentence', columns='Method', values='Avg_Accuracy')
            systems_vs_baseline_fce = ['dictionary', 'collocation', 'metalinguistic', 'mix']
            if 'baseline' in pivot_fce.columns:
                for system in systems_vs_baseline_fce:
                    if system in pivot_fce.columns:
                        diff = (pivot_fce['baseline'] - pivot_fce[system]).abs().dropna()
                        if not diff.empty:
                            add_selection(diff.idxmax(), f"系統差異最大 (FCE): baseline vs {system}", selected_sentences)

    if not master_ablation_df.empty:
        info_stats = master_ablation_df.dropna(subset=['Crucial Information Preservation']).groupby('key_sentence')['Crucial Information Preservation'].mean()
        if not info_stats.empty:
            add_selection(info_stats.idxmax(), "Ablation: 資訊保存度最佳", selected_sentences)
            add_selection(info_stats.idxmin(), "Ablation: 資訊保存度最差", selected_sentences)

        pref_df = master_ablation_df.dropna(subset=['Preference']).copy()
        len_comp_df = pref_df[pref_df['Ablation'] == 'length = 80'].copy()
        if not len_comp_df.empty:
            len_map = {'50': -1, 'A': -1, '80': 1, 'B': 1, 'C': 0}
            len_comp_df['pref_score'] = len_comp_df['Preference'].astype(str).map(len_map)
            len_pref_scores = len_comp_df.groupby('key_sentence')['pref_score'].sum()
            if not len_pref_scores.empty:
                add_selection(len_pref_scores.idxmax(), "Ablation: 明顯偏好 'length = 80'", selected_sentences)

        subseq_comp_df = pref_df[(pref_df['Ablation'] == '+subsequent') & (pref_df['dataset'] == 'fce')].copy()
        if not subseq_comp_df.empty:
            subseq_map = {'50': -1, 'A': -1, 'sub': 1, 'subseq': 1, 'B': 1, 'C': 0}
            subseq_comp_df['pref_score'] = subseq_comp_df['Preference'].astype(str).map(subseq_map)
            subseq_pref_scores = subseq_comp_df.groupby('key_sentence')['pref_score'].sum()
            if not subseq_pref_scores.empty:
                add_selection(subseq_pref_scores.idxmax(), "Ablation: 明顯偏好 '+subsequent'", selected_sentences)

    if not selected_sentences:
        print("未挑選出任何質性分析案例。")
        return

    if not master_method_df.empty:
        master_method_df.rename(columns={'Method': 'System'}, inplace=True)
    if not master_ablation_df.empty:
        master_ablation_df.rename(columns={'Ablation': 'System'}, inplace=True)

    combined_df = pd.concat([master_method_df, master_ablation_df], ignore_index=True, sort=False)

    qualitative_examples = []
    for key, reasons in selected_sentences.items():
        reason_str = "; ".join(sorted(reasons))
        full_data = combined_df[combined_df['key_sentence'] == key].copy()
        full_data['reason_for_selection'] = reason_str
        qualitative_examples.append(full_data)

    final_report_df = pd.concat(qualitative_examples).drop_duplicates()

    cols_order = ['reason_for_selection', 'key_sentence', 'dataset', 'evaluator', 'System', 'Content', 'Accuracy', 'Helpfulness', 'Preference', 'Crucial Information Preservation']
    final_report_df = final_report_df[[col for col in cols_order if col in final_report_df.columns]]

    try:
        with pd.ExcelWriter(output_filepath, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
            final_report_df.to_excel(writer, sheet_name='Qualitative_Analysis_Examples', index=False)
        print(f"質性分析案例已成功儲存至 '{output_filepath}' 的 'Qualitative_Analysis_Examples' 分頁。")
    except FileNotFoundError:
        print(f"錯誤：找不到目標檔案 '{output_filepath}'。請先執行 generate_summary_report()。")
    except Exception as e:
        print(f"儲存質性分析案例時發生錯誤：{e}")


def find_error_analysis_examples(all_method_data, all_ablation_data, output_filepath):
    """
    [**新功能**] 尋找每個 Method 和 Ablation 中，平均 Accuracy 最低的例子以進行錯誤分析。
    """
    print("\n--- 開始尋找錯誤分析案例 (最低 Accuracy) ---")

    if not all_method_data and not all_ablation_data:
        print("沒有可用於錯誤分析的資料。")
        return

    master_method_df = pd.DataFrame()
    if all_method_data:
        master_method_df = pd.concat(all_method_data, ignore_index=True).copy()
        master_method_df['Accuracy'] = pd.to_numeric(master_method_df['Accuracy'], errors='coerce')

    master_ablation_df = pd.DataFrame()
    if all_ablation_data:
        master_ablation_df = pd.concat(all_ablation_data, ignore_index=True).copy()
        master_ablation_df['Accuracy'] = pd.to_numeric(master_ablation_df['Accuracy'], errors='coerce')

    error_cases = []

    if not master_method_df.empty:
        method_avg_acc = master_method_df.groupby(['Method', 'key_sentence'])['Accuracy'].mean().reset_index()
        for method_name in method_avg_acc['Method'].unique():
            subset_df = method_avg_acc[method_avg_acc['Method'] == method_name]
            if not subset_df.empty:
                worst_case = subset_df.loc[subset_df['Accuracy'].idxmin()]
                error_cases.append({
                    'key_sentence': worst_case['key_sentence'],
                    'System': method_name,
                    'Reason': f"最低正確性案例 (Method: {method_name})"
                })

    if not master_ablation_df.empty:
        ablation_avg_acc = master_ablation_df.dropna(subset=['Accuracy']).groupby(['Ablation', 'key_sentence'])['Accuracy'].mean().reset_index()
        for ablation_name in ablation_avg_acc['Ablation'].unique():
            subset_df = ablation_avg_acc[ablation_avg_acc['Ablation'] == ablation_name]
            if not subset_df.empty:
                worst_case = subset_df.loc[subset_df['Accuracy'].idxmin()]
                error_cases.append({
                    'key_sentence': worst_case['key_sentence'],
                    'System': ablation_name,
                    'Reason': f"最低正確性案例 (Ablation: {ablation_name})"
                })

    if not error_cases:
        print("未找到任何可用於錯誤分析的案例。")
        return

    if not master_method_df.empty:
        master_method_df.rename(columns={'Method': 'System'}, inplace=True)
    if not master_ablation_df.empty:
        master_ablation_df.rename(columns={'Ablation': 'System'}, inplace=True)

    combined_df = pd.concat([master_method_df, master_ablation_df], ignore_index=True, sort=False)

    report_data = []
    for case in error_cases:
        full_data = combined_df[
            (combined_df['key_sentence'] == case['key_sentence']) &
            (combined_df['System'] == case['System'])
        ].copy()
        full_data['reason_for_selection'] = case['Reason']
        report_data.append(full_data)

    final_report_df = pd.concat(report_data).drop_duplicates()

    cols_order = ['reason_for_selection', 'key_sentence', 'dataset', 'evaluator', 'System', 'Content', 'Accuracy', 'Helpfulness', 'Preference', 'Crucial Information Preservation']
    final_report_df = final_report_df[[col for col in cols_order if col in final_report_df.columns]]

    try:
        with pd.ExcelWriter(output_filepath, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
            final_report_df.to_excel(writer, sheet_name='Error_Analysis_Lowest_Acc', index=False)
        print(f"錯誤分析案例已成功儲存至 '{output_filepath}' 的 'Error_Analysis_Lowest_Acc' 分頁。")
    except FileNotFoundError:
        print(f"錯誤：找不到目標檔案 '{output_filepath}'。請先執行 generate_summary_report()。")
    except Exception as e:
        print(f"儲存錯誤分析案例時發生錯誤：{e}")


if __name__ == "__main__":
    BASE_FOLDER = "./postprocess"
    SOURCE_FILE = os.path.join(BASE_FOLDER, "output", "randomized_multisheet_with_headers.xlsx")
    TARGET_FOLDER = os.path.join(BASE_FOLDER, "data", "evaluators")
    RESTORED_DATA_FOLDER = os.path.join(BASE_FOLDER, "data", "restored")
    os.makedirs(RESTORED_DATA_FOLDER, exist_ok=True)
    STATISTIC_FOLDER = os.path.join(BASE_FOLDER, "data", "statistics")
    os.makedirs(STATISTIC_FOLDER, exist_ok=True)

    print("\n--- 開始標頭還原流程 ---")
    target_files_to_restore = [f for f in os.listdir(TARGET_FOLDER) if f.endswith(".xlsx") and not f.startswith("~")]
    for file in target_files_to_restore:
        target_file_path = os.path.join(TARGET_FOLDER, file)
        output_file_path = os.path.join(RESTORED_DATA_FOLDER, file)
        print(f"\n正在處理 '{file}'...")
        restore_method_headers(
            source_headers_file=SOURCE_FILE,
            target_numbered_file=target_file_path,
            output_file=output_file_path
        )
    print("\n--- 標頭還原流程完成 ---")

    method_data, ablation_data = analyze_evaluation_results(RESTORED_DATA_FOLDER)

    summary_output_path = os.path.join(STATISTIC_FOLDER, "evaluation_analysis_summary.xlsx")
    generate_summary_report(method_data, ablation_data, summary_output_path)
    perform_additional_ablation_analysis(ablation_data, summary_output_path)
    analyze_by_sentence_category(method_data, summary_output_path)
    perform_t_tests(method_data, summary_output_path)
    perform_ablation_t_tests(method_data, ablation_data, summary_output_path)
    perform_preference_chi_squared_test(ablation_data, summary_output_path)
    perform_method_irr_analysis(method_data, summary_output_path)
    perform_ablation_irr_analysis(ablation_data, summary_output_path)
    select_qualitative_analysis_examples(method_data, ablation_data, summary_output_path)
    find_error_analysis_examples(method_data, ablation_data, summary_output_path)

    print("\n--- 分析工作流程已完成 ---")