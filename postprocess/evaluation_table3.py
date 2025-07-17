import pandas as pd
import random
import math
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font

# The source_data_blocks definition remains the same.
# 【NEW】I've added a "fce" block to demonstrate the new conditional logic.
source_data_blocks = [
    {
        "fixed_data": {
            "dataset": "longman",
            "index": 0,
            "formatted sentence": "It started pouring with rain and we all got {+soaked+}[-completely wet-]."
        },
        "random_data": [
            {
                "header": "baseline",
                "content_en": "The change was made because 'soaked' is a more concise, natural, and common way to express being very wet from rain, compared to 'completely wet,' which is more descriptive but less idiomatic.",
                "content_zh": "這個改變是因為「soaked」比「completely wet」更簡潔、自然且常用來描述被雨淋得很濕的情況；相較之下，「completely wet」較具描述性但較不地道。"
            },
            {
                "header": "dictionary",
                "content_en": "Replacing 'completely wet' with 'soaked' improves conciseness and naturalness. 'Soaked' is a common collocation with 'rain' and emphasizes being very wet, fitting idiomatic usage. It is more vivid and appropriate in informal contexts, enhancing clarity and impact.",
                "content_zh": "將「completely wet」替換為「soaked」提升了簡潔性和自然度。「Soaked」與「rain」常常搭配，強調非常濕，符合慣用語。它在非正式語境中更生動、恰當，增強了清晰度和表達力。"
            },
            {
                "header": "collocation",
                "content_en": "Replacing 'completely wet' with 'soaked' improves collocation naturalness, as 'soaked' is a common, vivid adjective used with 'wet' to emphasize being thoroughly drenched. 'Soaked' is more idiomatic and frequently co-occurs with 'wet' in corpus data, making the expression more concise and expressive.",
                "content_zh": "將「completely wet」替換為「soaked」提升了搭配的自然度，因為「soaked」是常用來生動形容詞，與「wet」搭配能更強調濕透。「soaked」在語料庫中與「wet」的共同出現頻率較高，使該表達更簡潔且富有表現力。"
            },
            {
                "header": "metalinguistic",
                "content_en": "Replacing 'completely wet' with 'soaked' improves conciseness and collocates better with 'pouring rain,' which is a common idiomatic expression. 'Soaked' is more vivid and frequently used in both spoken and written English to describe being very wet, fitting informal contexts.",
                "content_zh": "將「completely wet」替換為「soaked」提升了簡潔性，且與「pouring rain」這個慣用語搭配更自然。「soaked」更生動，常用於口語和書面語中，描述非常濕的狀態，適合非正式語境。"
            },
            {
                "header": "mix",
                "content_en": "Replacing 'completely wet' with 'soaked' improves collocation and naturalness, as 'soaked' is a common, vivid collocation that emphasizes being very wet. 'Soaked' is more idiomatic and frequently used in both spoken and written English to describe being thoroughly wet from rain.",
                "content_zh": "將「completely wet」替換為「soaked」改善了搭配的自然度，因為「soaked」是與「rain」常見且生動的搭配，強調非常濕。「soaked」在口語和書面語中都更地道，常用來描述被雨淋濕。"
            },
            {'header': 'original', 'content_en': 'lllll', 'content_zh': ''}
        ],
        'static_data': {
            'example_sentences': {
                'header': 'illustrative example:\nexample sentences',
                'content_en': '["After walking in the rain...", "After walking in the rain..."]',
                'content_zh': '["走在雨中後...", "走在雨中後..."]'
            },
            'example_collocations': {
                'header': 'illustrative example:\nexample collocations',
                'content_en': 'corresponding: ["wet -> soaked..."]...',
                'content_zh': 'ppp'
            },
            'output_80': {
                'header': 'output length = 80',
                'content_en': "Replacing 'completely wet' with 'soaked' improves...",
                'content_zh': "將「completely wet」替換為「soaked」提升了..."
            },
            'input_add_one': {
                'header': 'input sentence:\nadd one subsequent sentence',
                'content_en': '',
                'content_zh': ''
            },
            'output_50': {
                'header': 'output length = 50:\ninput sentence only',
                'content_en': "Replacing 'completely wet' with 'soaked' improves...",
                'content_zh': "將「completely wet」替換為「soaked」提升了..."
            }
        }
    },
    {
        "fixed_data": {
            "dataset": "fce",
            "index": 1,
            "formatted sentence": "To [-accomplish-]{+achieve+} world unity, we need peace."
        },
        "random_data": [
            {
                "header": "baseline",
                "content_en": "The change was made because 'achieve' is a more natural and commonly used verb with 'world unity' than 'accomplish.' 'Achieve' better fits the context of reaching a goal like unity.",
                "content_zh": "這個改動是因為「achieve」比「accomplish」更自然且常用於「world unity」這樣的語境。「achieve」更適合用來描述達成像團結這樣的目標。"
            },
            {
                "header": "dictionary",
                "content_en": "'Achieve' means to successfully reach a desired objective, especially after effort, and is often used with abstract nouns like 'unity.' 'Accomplish' is similar but less frequently collocates with 'unity.'",
                "content_zh": "「achieve」指的是經過努力後成功達成目標，常與「unity」等抽象名詞搭配。「accomplish」意思相近，但較少與「unity」連用。"
            },
            {
                "header": "collocation",
                "content_en": "'Achieve' is the standard collocation with 'unity' in English. Phrases like 'achieve unity' or 'achieve world unity' are common, while 'accomplish unity' is rarely used.",
                "content_zh": "「achieve」是與「unity」最常見的搭配，如「achieve unity」或「achieve world unity」在英文中很常見，而「accomplish unity」則很少見。"
            },
            {
                "header": "metalinguistic",
                "content_en": "The verb 'achieve' is more idiomatic when talking about abstract goals like 'world unity.' It emphasizes the process and effort needed to reach unity, which fits the intended meaning.",
                "content_zh": "當談論像「world unity」這樣的抽象目標時，動詞「achieve」更符合語感，強調達成團結所需的過程與努力，貼合原意。"
            },
            {
                "header": "mix",
                "content_en": "Replacing 'accomplish' with 'achieve' improves naturalness and collocation. 'Achieve world unity' is a common phrase, making the sentence sound more fluent and idiomatic.",
                "content_zh": "將「accomplish」改為「achieve」提升了自然度和搭配性。「achieve world unity」是常見用法，使句子更流暢、地道。"
            },
            {'header': 'original', 'content_en': '', 'content_zh': ''}
        ],
        'static_data': {
            'example_sentences': {
                'header': 'illustrative example:\nexample sentences',
                'content_en': 'rrr',
                'content_zh': 'rrr'
            },
            'example_collocations': {
                'header': 'illustrative example:\nexample collocations',
                'content_en': 'aaa',
                'content_zh': 'aaa'
            },
            'output_80': {
                'header': 'output length = 80',
                'content_en': "bbb",
                'content_zh': "bbb"
            },
            'input_add_one': {
                'header': 'input sentence:\nadd one subsequent sentence',
                'content_en': 'ccc',
                'content_zh': 'ccc'
            },
            'output_50': {
                'header': 'output length = 50:\ninput sentence only',
                'content_en': "ddd",
                'content_zh': "ddd"
            }
        }
    }
]


def generate_evaluation_table3(source_data_blocks):
    # 1. ===== Define output filenames =====
    output_filename_with_headers = "./postprocess/output/randomized_multisheet_with_headers.xlsx"
    output_filename_special_header = "./postprocess/output/randomized_multisheet_no_headers.xlsx"

    # --- Define the mapping for static data headers ---
    static_header_map = {
        "illustrative example:\nexample sentences": "sentences",
        "illustrative example:\nexample collocations": "collocations",
        "output length = 80": "length = 80",
        "input sentence:\nadd one subsequent sentence": "+subsequent",
        "output length = 50:\ninput sentence only": "length = 50"
    }

    # ---【Step 1: Use Pandas to Write Data】---
    print("Step 1: Writing data with Pandas...")
    with pd.ExcelWriter(output_filename_with_headers, engine='openpyxl') as writer_regular, \
         pd.ExcelWriter(output_filename_special_header, engine='openpyxl') as writer_special:

        for block in source_data_blocks:
            sheet_name = f"{block['fixed_data']['dataset']}_{block['fixed_data']['index']}"
            print(f"  - Processing sheet: {sheet_name}")

            random.shuffle(block['random_data'])
            random_cols = ['Method', 'Content', 'Accuracy', 'Helpfulness']

            data_for_df_random_regular = []
            for item in block['random_data']:
                data_for_df_random_regular.append([item['header'], item['content_en'], '', ''])
                data_for_df_random_regular.append(['', item['content_zh'], '', ''])
            df_random_regular = pd.DataFrame(data_for_df_random_regular, columns=random_cols)

            data_for_df_random_special = []
            method_counter = 1
            for item in block['random_data']:
                data_for_df_random_special.append([str(method_counter), item['content_en'], '', ''])
                data_for_df_random_special.append(['', item['content_zh'], '', ''])
                method_counter += 1
            df_random_special = pd.DataFrame(data_for_df_random_special, columns=random_cols)

            static_keys_ordered = list(block['static_data'].keys())
            data_for_df_static = []
            static_cols = ['Ablation Study', 'Content', 'Accuracy', 'Helpfulness', 'Preference', 'Crucial Information Preservation']

            for key in static_keys_ordered:
                item = block['static_data'][key]
                original_header = item['header']
                display_header = static_header_map.get(original_header, original_header)
                data_for_df_static.append([display_header, item['content_en'], '', '', '', ''])
                data_for_df_static.append(['', item['content_zh'], '', '', '', ''])
            df_static = pd.DataFrame(data_for_df_static, columns=static_cols)

            start_row_data = 2
            start_row_static = df_random_regular.shape[0] + 4

            df_random_regular.to_excel(writer_regular, sheet_name=sheet_name, startrow=start_row_data, index=False, header=False)
            df_static.to_excel(writer_regular, sheet_name=sheet_name, startrow=start_row_static, index=False, header=False)

            df_random_special.to_excel(writer_special, sheet_name=sheet_name, startrow=start_row_data, index=False, header=False)
            df_static.to_excel(writer_special, sheet_name=sheet_name, startrow=start_row_static, index=False, header=False)

    print("  - Data writing complete for both files.")

    # ---【Step 2: Define the Reusable Formatting Function】---
    def apply_formatting(workbook, source_data_blocks):
        print(f"\nApplying formatting to workbook...")

        # --- 【MODIFIED】Define styles for fill and border ---
        light_grey_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        wrap_alignment_style = Alignment(wrap_text=True, vertical='top', horizontal='left')
        large_header_font = Font(size=22)
        regular_font = Font(size=18)
        center_align_style = Alignment(horizontal='center', vertical='center')

        for block in source_data_blocks:
            sheet_name = f"{block['fixed_data']['dataset']}_{block['fixed_data']['index']}"
            sheet = workbook[sheet_name]
            print(f"  - Formatting sheet: {sheet_name}")

            for row in sheet.iter_rows(min_row=3):
                for cell in row:
                    cell.font = regular_font

            # Define layout dimensions
            num_random_rows = len(block['random_data']) * 2
            static_keys_ordered = list(block['static_data'].keys())

            SENTENCE_ROW, HEADER_ROW = 1, 2
            RANDOM_CONTENT_START_ROW = 3
            RANDOM_CONTENT_END_ROW = HEADER_ROW + num_random_rows
            STATIC_HEADER_ROW = RANDOM_CONTENT_END_ROW + 2
            STATIC_CONTENT_START_ROW = STATIC_HEADER_ROW + 1
            STATIC_CONTENT_END_ROW = STATIC_HEADER_ROW + len(static_keys_ordered) * 2

            num_random_cols, num_static_cols = 4, 6
            max_cols = max(num_random_cols, num_static_cols)

            # Set column widths
            content_col_width = 130
            sheet.column_dimensions['A'].width = 18
            sheet.column_dimensions['B'].width = content_col_width
            sheet.column_dimensions['C'].width = 10
            sheet.column_dimensions['D'].width = 10
            sheet.column_dimensions['E'].width = 10
            sheet.column_dimensions['F'].width = 10

            # Format Sentence and Header Rows
            sheet.merge_cells(start_row=SENTENCE_ROW, start_column=1, end_row=SENTENCE_ROW, end_column=max_cols)
            sentence_cell = sheet.cell(row=SENTENCE_ROW, column=1)
            sentence_cell.value = block['fixed_data']['formatted sentence']
            sentence_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            sentence_cell.font = large_header_font

            headers_rand = ['Method', 'Content', 'Acc', 'Helpful']
            for c, val in enumerate(headers_rand, 1): sheet.cell(row=HEADER_ROW, column=c).value = val
            headers_static = ['Ablation', 'Content', 'Acc', 'Helpful', 'Pref', 'Info']
            for c, val in enumerate(headers_static, 1): sheet.cell(row=STATIC_HEADER_ROW, column=c).value = val

            for r in [HEADER_ROW, STATIC_HEADER_ROW]:
                sheet.row_dimensions[r].height = 20
                for c in range(1, max_cols + 1):
                    cell = sheet.cell(row=r, column=c)
                    if cell.value:
                        cell.alignment = center_align_style
                        cell.font = Font(bold=True, size=15)

            all_content_rows_indices = [SENTENCE_ROW] + \
                list(range(RANDOM_CONTENT_START_ROW, RANDOM_CONTENT_END_ROW + 1)) + \
                list(range(STATIC_CONTENT_START_ROW, STATIC_CONTENT_END_ROW + 1))
            for row_num in all_content_rows_indices:
                sheet.cell(row=row_num, column=1).fill = light_grey_fill
                sheet.cell(row=row_num, column=2).alignment = wrap_alignment_style

            # Dynamic Row Height Calculation
            CHARS_PER_LINE_ESTIMATE = content_col_width * 1.3
            HEIGHT_PER_LINE = 18
            MIN_ROW_HEIGHT = 40

            for row_num in all_content_rows_indices:
                content_col_idx = 1 if row_num == SENTENCE_ROW else 2
                cell = sheet.cell(row=row_num, column=content_col_idx)

                if isinstance(cell.value, str) and cell.value:
                    effective_len = sum(2 if '\u4e00' <= char <= '\u9fff' else 1 for char in cell.value)
                    num_lines = math.ceil(effective_len / CHARS_PER_LINE_ESTIMATE) + 1
                    estimated_height = (num_lines * HEIGHT_PER_LINE) + 15
                    sheet.row_dimensions[row_num].height = max(MIN_ROW_HEIGHT, estimated_height)
                else:
                    sheet.row_dimensions[row_num].height = MIN_ROW_HEIGHT

            # --- 【MODIFIED】Conditional Cell Filling & Bordering ---
            acc_col, help_col, pref_col, cruc_col = 3, 4, 5, 6
            dataset_type = block['fixed_data']['dataset']
            shuffled_random_data = block['random_data']

            # --- Loop 1: Random data section ---
            for i, item in enumerate(shuffled_random_data):
                r_en = RANDOM_CONTENT_START_ROW + i * 2 # This is the English row
                should_fill = (dataset_type == 'longman') or \
                              (dataset_type == 'fce' and item['header'] != 'original')
                if should_fill:
                    for c in [acc_col, help_col]:
                        cell = sheet.cell(row=r_en, column=c)
                        cell.fill = light_grey_fill
                        cell.border = thin_border

            # --- Loop 2: Static data section ---
            for i, key in enumerate(static_keys_ordered):
                r_en = STATIC_CONTENT_START_ROW + i * 2 # This is the English row
                header_text = block['static_data'][key]['header']
                cols_to_fill = []
                if header_text in ['illustrative example:\nexample sentences', 'illustrative example:\nexample collocations']:
                    cols_to_fill = [acc_col, help_col]
                elif header_text == 'output length = 80':
                    cols_to_fill = [acc_col, help_col, pref_col, cruc_col]
                elif header_text == 'input sentence:\nadd one subsequent sentence':
                    cols_to_fill = [acc_col, help_col, pref_col]

                for c in cols_to_fill:
                    cell = sheet.cell(row=r_en, column=c)
                    cell.fill = light_grey_fill
                    cell.border = thin_border

    # ---【Step 3: Apply Formatting to Both Files】---
    print("\n--- Starting File Generation and Formatting ---")
    workbook_special = openpyxl.load_workbook(output_filename_special_header)
    apply_formatting(workbook_special, source_data_blocks)
    workbook_special.save(output_filename_special_header)
    print(f"✅ Successfully formatted and saved '{output_filename_special_header}'")

    workbook_regular = openpyxl.load_workbook(output_filename_with_headers)
    apply_formatting(workbook_regular, source_data_blocks)
    workbook_regular.save(output_filename_with_headers)
    print(f"✅ Successfully formatted and saved '{output_filename_with_headers}'")

    print(f"\n✅ All tasks completed successfully!")


if __name__ == "__main__":
    # In a real scenario, you would import your source data like this:
    from .source_data_blocks import source_data_blocks
    generate_evaluation_table3(source_data_blocks)
    print("\nEvaluation table generation complete.")
